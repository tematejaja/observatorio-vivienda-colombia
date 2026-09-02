# -*- coding: utf-8 -*-
"""
FASE 2 - Paso 3: integrar el deficit habitacional en la tabla maestra.

Reemplaza los marcadores `ND - pendiente Fase 2 (ECV)` del bloque
`deficit_habitacional` por las estimaciones reales de `42_deficit_ciudades.py`.

Tres decisiones que conviene tener a la vista:

  * Se publican TODOS los indicadores calculados, incluidos aquellos que en la
    mayoria de ciudades salen etiquetados NO PUBLICAR. Ocultarlos daria la
    impresion de que la ECV puede sostener afirmaciones por ciudad que en
    realidad no puede sostener; la etiqueta de confiabilidad existe justamente
    para decirlo. El unico deficit que aguanta lectura por ciudad es el TOTAL y
    el CUALITATIVO; el CUANTITATIVO es tan poco frecuente en cabecera (mediana
    2,2%) que la muestra de la ECV no lo estima con precision util.

  * 2026 queda en ND: la ECV 2026 todavia no ha sido publicada. Es el mismo
    criterio que ya se aplica a Pobreza Monetaria 2026, y evita extrapolar.

  * El libro de Excel tiene hojas derivadas de la Fase 1 que este paso no
    recalcula; solo se actualiza la hoja `Deficit_Habitacional`. El CSV maestro
    es la fuente que lee la aplicacion (`datos_observatorio.py`).
"""
import subprocess
import sys
from pathlib import Path

import pandas as pd

SCRIPTS_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPTS_DIR))
from config_ciudades import stdout_utf8

BASE_DIR = SCRIPTS_DIR.parent
MAESTRA_CSV = BASE_DIR / "output" / "observatorio_vivienda_capitales_2023_2026.csv"
MAESTRA_XLSX = BASE_DIR / "output" / "observatorio_vivienda_capitales_2023_2026.xlsx"
DEFICIT_CSV = BASE_DIR / "GEIH" / "procesado_nacional" / "deficit_23ciudades.csv"

BLOQUE = "deficit_habitacional"
CATALOGO_ANDA = {2023: 827, 2024: 861, 2025: 905}
ANIO_SIN_DATO = "2026*"

NOTA = ("Metodologia DANE de deficit habitacional (2020), criterios de cabecera; "
        "replica validada contra el anexo oficial de la ECV")
NOTA_2026 = "ND - el DANE aun no publica la ECV 2026; no se extrapola"


def fmt(v, dec):
    return "" if pd.isna(v) else f"{float(v):.{dec}f}"


def construir_filas(maestra: pd.DataFrame, det: pd.DataFrame) -> pd.DataFrame:
    dominio = dict(maestra[["ciudad", "codigo_dominio"]].drop_duplicates().values)
    indicadores = sorted(det["indicador"].unique())
    filas = []

    for _, r in det.iterrows():
        anio = int(r["anio"])
        filas.append({
            "ciudad": r["ciudad"], "codigo_dominio": dominio[r["ciudad"]],
            "anio": str(anio), "meses_incluidos": "Ene-Dic (12 meses)",
            "bloque_indicador": BLOQUE, "nombre_indicador": r["indicador"],
            "valor": fmt(r["valor_pct"], 2),
            "error_estandar": fmt(r["error_estandar"], 4),
            "ic95_inf": fmt(r["ic95_inf"], 4), "ic95_sup": fmt(r["ic95_sup"], 4),
            "deff": fmt(r["deff"], 4), "n_muestral": str(int(r["n_muestral"])),
            "cv_pct": fmt(r["cv_pct"], 4),
            "etiqueta_confiabilidad": r["etiqueta_confiabilidad"],
            "fuente": f"ECV {anio} (DANE, catalogo ANDA {CATALOGO_ANDA[anio]})",
            "observacion": (f"{NOTA}. {r['nota']}"
                            if pd.notna(r.get("nota")) and str(r["nota"]).strip()
                            else NOTA),
        })

    for ciudad in sorted(det["ciudad"].unique()):
        for nombre in indicadores:
            filas.append({
                "ciudad": ciudad, "codigo_dominio": dominio[ciudad],
                "anio": ANIO_SIN_DATO, "meses_incluidos": "Ene-Dic (12 meses)",
                "bloque_indicador": BLOQUE, "nombre_indicador": nombre,
                "valor": "ND - ECV 2026 no publicada", "error_estandar": "",
                "ic95_inf": "", "ic95_sup": "", "deff": "", "n_muestral": "0",
                "cv_pct": "", "etiqueta_confiabilidad": "ND",
                "fuente": "ND - DANE no publica ECV 2026",
                "observacion": NOTA_2026,
            })
    return pd.DataFrame(filas)


def main() -> None:
    stdout_utf8()
    print("FASE 2 · Paso 3 — integración del déficit en la tabla maestra")
    print("=" * 74)

    maestra = pd.read_csv(MAESTRA_CSV, dtype=str, keep_default_na=False)
    det = pd.read_csv(DEFICIT_CSV)

    orden = ["ciudad", "anio", "bloque_indicador", "nombre_indicador"]
    if not maestra[orden].equals(maestra.sort_values(orden, kind="stable")[orden].reset_index(drop=True)):
        raise RuntimeError("La tabla maestra no está ordenada como se asumía; "
                           "revisar antes de reescribirla.")

    antes = len(maestra)
    n_placeholder = int((maestra["bloque_indicador"] == BLOQUE).sum())
    resto = maestra[maestra["bloque_indicador"] != BLOQUE]
    nuevas = construir_filas(maestra, det)

    if set(nuevas.columns) != set(maestra.columns):
        raise RuntimeError(f"Columnas distintas: {set(nuevas.columns) ^ set(maestra.columns)}")

    salida = (pd.concat([resto, nuevas[maestra.columns]], ignore_index=True)
              .sort_values(orden, kind="stable").reset_index(drop=True))
    salida.to_csv(MAESTRA_CSV, index=False, encoding="utf-8-sig")

    print(f"  filas: {antes:,} -> {len(salida):,}")
    print(f"  bloque '{BLOQUE}': {n_placeholder} marcadores ND -> {len(nuevas):,} filas")
    print(f"  indicadores nuevos: {det['indicador'].nunique()} × 23 ciudades × 4 años")

    pub = nuevas[nuevas["anio"] != ANIO_SIN_DATO]
    print("\n  confiabilidad de las filas con dato:")
    for etiqueta, k in pub["etiqueta_confiabilidad"].value_counts().items():
        print(f"    {etiqueta:<14}{k:>5}")

    # Hoja de Excel del bloque (las demas hojas son artefactos de la Fase 1).
    hoja = pub[["ciudad", "anio", "nombre_indicador", "valor", "fuente", "observacion"]]
    import openpyxl                                   # noqa: PLC0415
    wb = openpyxl.load_workbook(MAESTRA_XLSX)
    if "Deficit_Habitacional" in wb.sheetnames:
        del wb["Deficit_Habitacional"]
    ws = wb.create_sheet("Deficit_Habitacional")
    ws.append(list(hoja.columns))
    for fila in hoja.itertuples(index=False):
        ws.append(list(fila))
    wb.save(MAESTRA_XLSX)
    print(f"\n  hoja 'Deficit_Habitacional' del libro reescrita ({len(hoja):,} filas)")

    print("\n  regenerando fichas de ciudad...")
    r = subprocess.run([sys.executable, str(SCRIPTS_DIR / "30_generar_fichas_ciudades.py")],
                       capture_output=True, text=True, encoding="utf-8", errors="replace")
    if r.returncode != 0:
        print(r.stdout[-1500:])
        print(r.stderr[-1500:])
        raise SystemExit("Falló la regeneración de fichas")
    print("  " + (r.stdout.strip().splitlines() or ["(sin salida)"])[-1])


if __name__ == "__main__":
    main()
